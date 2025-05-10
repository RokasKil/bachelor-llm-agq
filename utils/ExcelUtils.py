
from typing import Tuple
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Alignment
from openpyxl.styles import Border, Side
from openpyxl.cell import Cell, MergedCell
from copy import copy
def write_row(worksheet: Worksheet, values: list, row: int, col: int = 1):
    for index, value in enumerate(values):
        worksheet.cell(row=row, column=col + index, value=str(value))

def apply_border(cell: Cell, border: Border):
    joint_border = copy(cell.border)
    if border.top != None:
        joint_border.top = border.top
    if border.left != None:
        joint_border.left = border.left
    if border.right != None:
        joint_border.right = border.right
    if border.bottom != None:
        joint_border.bottom = border.bottom
    cell.border = joint_border
    
def set_boundary_border(worksheet: Worksheet, min_col: int, min_row: int, max_col: int, max_row: int, border: Border = None):
    if border is None:
        border = Border(
            left=Side(style='medium'),
            right=Side(style='medium'),
            top=Side(style='medium'),
            bottom=Side(style='medium')
        )

    for row in worksheet.iter_rows(min_row=min_row, max_row=min_row, min_col=min_col, max_col=max_col):
        for cell in row:
            apply_border(cell, Border(top=border.top)) # Only top border

    for row in worksheet.iter_rows(min_row=max_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            apply_border(cell, Border(bottom=border.bottom)) # Only bottom border

    for row in worksheet.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=min_col):
        for cell in row:
            apply_border(cell, Border(left=border.left)) # Only left border

    for row in worksheet.iter_rows(min_row=min_row, max_row=max_row, min_col=max_col, max_col=max_col):
        for cell in row:
            apply_border(cell, Border(right=border.right)) # Only right border

    # Apply corner borders
    apply_border(worksheet.cell(row=min_row, column=min_col), Border(top=border.top, left=border.left))
    apply_border(worksheet.cell(row=min_row, column=max_col), Border(top=border.top, right=border.right))
    apply_border(worksheet.cell(row=max_row, column=min_col), Border(bottom=border.bottom, left=border.left))
    apply_border(worksheet.cell(row=max_row, column=max_col), Border(bottom=border.bottom, right=border.right))


def set_grid_border(worksheet: Worksheet, min_col: int, min_row: int, max_col: int, max_row: int, border: Border = None):
    if border is None:
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

    for row in worksheet.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            apply_border(cell, border)

def set_combined_border(worksheet: Worksheet, min_col: int, min_row: int, max_col: int, max_row: int, grid_border: Border = None, boundary_border: Border = None):
    set_grid_border(worksheet, min_col, min_row, max_col, max_row, grid_border)
    set_boundary_border(worksheet, min_col, min_row, max_col, max_row, boundary_border)

def get_questionnaire_info(worksheet: Worksheet) -> Tuple[str, str, str, int]:
    method_key = worksheet.cell(1, 2).value
    category = worksheet.cell(2, 2).value
    file = worksheet.cell(3, 2).value
    run_index = int(worksheet.cell(4, 2).value)
    
    return (category, file, run_index, method_key)

def get_questionnaire(worksheet: Worksheet) -> Tuple[dict, int]:
    # there should only be one type of merged cells so doesn't matter which one we take
    if worksheet.merged_cells and len(worksheet.merged_cells.ranges) > 0:
        merged_range = next(iter(worksheet.merged_cells.ranges))
        answers_count = merged_range.max_row - merged_range.min_row + 1
    else:
        answers_count = 1
    questions = []
    index = 0
    while True:
        row_index = 7 + index * answers_count
        if worksheet.cell(row_index, 1).value != None:
            questions.append({
                "question_text": worksheet.cell(row_index, 1).value,
                "answer": worksheet.cell(row_index, 2).value
            })
            index += 1
        else:
            break
    questionnaire = {"questions": questions}
    return (questionnaire, answers_count)