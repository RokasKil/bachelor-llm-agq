from dotenv import load_dotenv
load_dotenv()

import os
from openpyxl import load_workbook
from utils.ExcelUtils import *
from methods.Config import METHODS
from enums.AppendBehaviour import AppendBehavior


EXCEL_FILE = os.getenv("EXCEL_OUT", "data.xlsx")
SOURCES_PATH = os.getenv("SOURCES_PATH", "sources")
APPEND_BEHAVIOUR = AppendBehavior[os.getenv("EXCEL_APPEND_EXISTING_BEHAVIOUR", "SKIP")]
STARTING_ROW = 7
wb = load_workbook(EXCEL_FILE)
for ws in wb:
    if not ws.title.isnumeric():
        continue
    key = get_questionnaire_info(ws)
    print(f"Answering {key}")
    (category, file, run_index, method_key) = key
    file_path = os.path.join(SOURCES_PATH, category, file)
    if not os.path.exists(file_path):
        print(f"File doesn't exist {file_path}")
        continue
    with open(file_path, 'r', encoding='utf-8') as f:
        input_text = f.read()
    if method_key not in METHODS:
        print(f"Method not found: {method_key}")
        continue
    method = METHODS[method_key].method
    (questionnaire, answers_count) = get_questionnaire(ws)
    for answer_set_index in range(answers_count):
        if ws.cell(STARTING_ROW + answer_set_index, 9).value != None and APPEND_BEHAVIOUR == AppendBehavior.SKIP:
            print(f"Skipping {answer_set_index + 1} answer set (answers already evaluated)")
            continue
        answers: list[str] = []
        for index in range(len(questionnaire["questions"])):
            row_index = STARTING_ROW + index * answers_count + answer_set_index
            answers.append(str(ws.cell(row_index, 7).value or ''))
        if all(map(lambda x: x.strip() == "", answers)):
            print(f"Skipping {answer_set_index + 1} answer set (no answers provided)")
            continue
        print(f"Answering {answer_set_index + 1} answer set")
        method.reset_usage()
        evaluations = method.evaluate_answers(input_text, questionnaire, answers)
        ws.cell(2, 6 + answer_set_index, method.usage["completion_tokens"])
        ws.cell(3, 6 + answer_set_index, method.usage["prompt_tokens"])
        ws.cell(4, 6 + answer_set_index, method.usage["total_tokens"])
        for index, evaluation in enumerate(evaluations["evaluations"]):
            row_index = STARTING_ROW + index * answers_count + answer_set_index
            ws.cell(row_index, 9, evaluation["evaluation"])
        wb.save(EXCEL_FILE)