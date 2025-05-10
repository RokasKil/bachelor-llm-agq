from dotenv import load_dotenv
load_dotenv()

import os
from openpyxl import  load_workbook
from utils.ExcelUtils import *
from enums.AppendBehaviour import AppendBehavior
import random
from methods.AnswerGenerator import AnswerGenerator

EXCEL_FILE = os.getenv("EXCEL_OUT", "data.xlsx")
SOURCES_PATH = os.getenv("SOURCES_PATH", "sources")
APPEND_BEHAVIOUR = AppendBehavior[os.getenv("EXCEL_APPEND_EXISTING_BEHAVIOUR", "SKIP")]
STARTING_ROW = 7
method = AnswerGenerator()
wb = load_workbook(EXCEL_FILE)
for ws in wb:
    if not ws.title.isnumeric():
        continue

    key = get_questionnaire_info(ws)
    (questionnaire, answers_count) = get_questionnaire(ws)
    question_count = len(questionnaire["questions"])

    if all(ws.cell(row_index, 8).value != None for row_index in range(STARTING_ROW, STARTING_ROW + answers_count * question_count)) and APPEND_BEHAVIOUR == AppendBehavior.SKIP:
        print(f"Skipping {key}")
        continue
    print(f"Generating answers for {key}")
    for question in questionnaire["questions"]:
        generate_answers = ["correct", "incorrect", "incorrect"]
        if answers_count == 1:
            question["generate_answers"] = [random.choice(["correct", "incorrect"])]
        else:
            if answers_count > len(generate_answers):
                generate_answers += random.choices(["correct", "incorrect"], k = answers_count - len(generate_answers))
            random.shuffle(generate_answers)
            question["generate_answers"] = generate_answers
    answers = method.generate_answers(questionnaire["questions"], answers_count)
    for i, answer in enumerate(answer for answer_set in answers["answers"] for answer in answer_set):
        print (i, answer)
        ws.cell(STARTING_ROW + i, 7, answer["answer"])
        ws.cell(STARTING_ROW + i, 8, answer["type"])
    wb.save(EXCEL_FILE)