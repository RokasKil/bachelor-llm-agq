from dotenv import load_dotenv
load_dotenv()

import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from utils.ExcelUtils import *
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.workbook.defined_name import DefinedName
from methods.Config import METHODS
from enums.AppendBehaviour import AppendBehavior

EXCEL_FILE = os.getenv("EXCEL_OUT", "data.xlsx")
APPEND_FILE = os.getenv("EXCEL_APPEND", "false").lower() == "true"
APPEND_BEHAVIOUR = AppendBehavior[os.getenv("EXCEL_APPEND_EXISTING_BEHAVIOUR", "SKIP")]
QUESTIONNAIRE_COUNT_PER_MODEL = int(os.getenv("QUESTIONNAIRE_COUNT_PER_MODEL", "3"))
ANSWER_COUNT_PER_QUESTIONNAIRE= int(os.getenv("ANSWER_COUNT_PER_QUESTIONNAIRE", "3"))
SOURCES_PATH = os.getenv("SOURCES_PATH", "sources")


existing_sheets = {}
index = 0
if APPEND_FILE and os.path.exists(EXCEL_FILE):
    wb = load_workbook(EXCEL_FILE)
    for ws in wb:
        if ws.title.isnumeric():
            index = max(int(ws.title), index)
            existing_sheets[get_questionnaire_info(ws)] = ws.title
else:
    wb = Workbook()
    del wb["Sheet"]
    options_ws = wb.create_sheet("Options")
    # Define drop downs 
    write_row(options_ws, ["yes", "no"], row = 1)
    write_row(options_ws, ["correct", "incorrect", "unanswerable"], row = 2)
    wb.defined_names.add(DefinedName("binary_options", attr_text="Options!$A$1:$B$1"))
    wb.defined_names.add(DefinedName("answer_evaluations", attr_text="Options!$A$2:$C$2"))
    options_ws.sheet_state = "hidden"


for root, dirs, files in os.walk(SOURCES_PATH):
    for file in files:
        file_path = os.path.join(root, file)
        category = root[len(SOURCES_PATH + os.sep):]
        print(f"Found file: {file_path} in {category}")
        with open(file_path, 'r', encoding='utf-8') as f:
            input_text = f.read()
        for method_key, method_config in METHODS.items():
            if not method_config.use:
                continue
            method = method_config.method
            print("running agianst", method_key, method.__class__.__name__)
            for run_index in range(1, QUESTIONNAIRE_COUNT_PER_MODEL + 1):
                key = (category, file, run_index, method_key)
                if key in existing_sheets:
                    sheet_key = existing_sheets[key]
                    if APPEND_BEHAVIOUR == AppendBehavior.REPLACE:
                        print(f"replacing {key}")
                        old_index = wb.index(wb[sheet_key])
                        del wb[sheet_key]
                        wb.create_sheet(sheet_key, index=old_index)
                    else:
                        print(f"skipping {key}")
                        continue
                else:
                    print(f"creating {key}")
                    index += 1
                    sheet_key = str(index)
                    wb.create_sheet(sheet_key)
                wb.active = wb[sheet_key]
                ws = wb.active
                # Set column widths
                for i, width in enumerate([50, 20, 12, 12, 12, 20, 20, 15, 15]):
                    ws.column_dimensions[get_column_letter(i + 1)].width = width

                # Write data about questionnaire and keep one row as a spacer
                write_row(ws, ["Method", method_key], 1)
                write_row(ws, ["Category", category], 2)
                write_row(ws, ["File", file], 3)
                write_row(ws, ["Run", run_index], 4)
                set_combined_border(ws, 1, 1, 2, 4)

                write_row(ws, ["Tokens", "Questionnaire"] + [f"Answer set {i}" for i in range(1, ANSWER_COUNT_PER_QUESTIONNAIRE + 1)], 1, 4)
                write_row(ws, ["Completion"], 2, 4)
                write_row(ws, ["Prompt"], 3, 4)
                write_row(ws, ["Total"], 4, 4)
                set_combined_border(ws, 4, 1, 5 + ANSWER_COUNT_PER_QUESTIONNAIRE, 4)

                # Create DataValidation for dropdowns
                binary_dropdown = DataValidation(type="list", formula1="binary_options", allow_blank=True)
                answer_eval_dropdown = DataValidation(type="list", formula1="answer_evaluations", allow_blank=True)
                ws.add_data_validation(binary_dropdown)
                ws.add_data_validation(answer_eval_dropdown)
                # Table headers
                write_row(ws, ["Question", "Generated answer", "Relevant", "Clear", "Answerable", "Question Evaluation", "Answer", "True evaluation", "Evaluation"], 6)
                set_combined_border(ws, 1, 6, 9, 6)
                
                method.reset_usage()
                questionnaire = method.generate_questionnaire(input_text)
                ws.cell(2, 5, method.usage["completion_tokens"])
                ws.cell(3, 5, method.usage["prompt_tokens"])
                ws.cell(4, 5, method.usage["total_tokens"])
                row_index = 7
                # Write questions
                for question in questionnaire["questions"]:
                    # Merge question and answer cells to allow us to enter multiple answers for one question
                    for i in range(1, 7):
                        ws.merge_cells(start_row=row_index, start_column=i, end_row=row_index + ANSWER_COUNT_PER_QUESTIONNAIRE - 1, end_column=i)
                        ws.cell(row=row_index, column=i).alignment = Alignment(wrap_text=True, vertical="top")
                    # set dropdowns for question eval
                    for i in range(3, 6):
                        binary_dropdown.add(ws.cell(row=row_index, column=i))
                    # set question eval formula
                    ws.cell(row=row_index, column=6, value=f'=IF(COUNTIF(C{row_index}:E{row_index}, "yes") = 3, "yes", "no")')
                    # set dropdowns for answer eval
                    for row in ws.iter_rows(min_col=8, min_row=row_index, max_col=9, max_row=row_index + ANSWER_COUNT_PER_QUESTIONNAIRE  - 1):
                        for cell in row:
                            answer_eval_dropdown.add(cell)
                    write_row(ws, [question["question_text"], question["answer"]], row_index)
                    set_combined_border(ws, 1, row_index, 9, row_index + ANSWER_COUNT_PER_QUESTIONNAIRE - 1)
                    row_index += ANSWER_COUNT_PER_QUESTIONNAIRE
                wb.save(EXCEL_FILE)
            
wb.save(EXCEL_FILE)